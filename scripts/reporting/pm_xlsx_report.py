#!/usr/bin/env python3
"""
pm_xlsx_report.py

pm.db から 4 シートの Excel ブック (アクションアイテム / 決定事項 /
要注意事項 / プロジェクトの現在地) を生成し、Box にアップロードして
Slack Canvas にリンク + サマリの目録を投稿する。

設定: data/argus_config.yaml の report: セクション
    report:
      canvas_id: <CANVAS_ID>          # 目録投稿先
      box_folder_id: "123456789"       # XLSX のアップロード先 Box フォルダ
      filename: "pm_report.xlsx"       # 任意。固定名で版管理する

Usage:
    python3 scripts/pm_xlsx_report.py
    python3 scripts/pm_xlsx_report.py --since 2026-03-01
    python3 scripts/pm_xlsx_report.py --output report.xlsx --skip-upload --skip-canvas
"""
from __future__ import annotations

import argparse
import json
import re
import subprocess
import sys
from datetime import date
from pathlib import Path

import yaml
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.worksheet.worksheet import Worksheet

sys.path.insert(0, str(Path(__file__).resolve().parent.parent))
from canvas_utils import post_to_canvas, sanitize_for_canvas
from cli_utils import (
    add_dry_run_arg,
    add_filter_arg,
    add_no_encrypt_arg,
    add_since_arg,
    make_logger,
    resolve_filter_presets,
    resolve_report_canvas_id,
)
from db_utils import fetch_milestone_progress, open_db, open_pm_db
from pm_report import (
    detect_risk_items,
    fetch_open_action_items,
    fetch_recent_decisions,
)

REPO_ROOT = Path(__file__).resolve().parent.parent.parent
DEFAULT_PM_DB = REPO_ROOT / "data" / "pm.db"
DEFAULT_CONFIG = REPO_ROOT / "data" / "argus_config.yaml"
DEFAULT_FILENAME = "pm_report.xlsx"

SHEET_AI = "アクションアイテム"
SHEET_DEC = "決定事項"
SHEET_RISK = "要注意事項"
SHEET_MS = "プロジェクトの現在地"
SHEET_ACH = "実績"

# 編集可能フィールド（pm_xlsx_sync.py と同じ定義を共有する）
AI_EDITABLE = ["content", "assignee", "due_date", "milestone_id", "status", "note"]
DEC_EDITABLE = ["content", "decided_at", "acknowledged"]

# シート列定義: (列名, DB列, 編集可)
AI_COLUMNS = [
    ("削除",         "_delete_flag",      True),
    ("id",           "id",                False),
    ("発生日",       "_extracted_date",   False),
    ("内容",         "content",           True),
    ("担当者",       "assignee",          True),
    ("期限",         "due_date",          True),
    ("MS",           "milestone_id",      True),
    ("状況",         "status",            True),
    ("対応状況",     "note",              True),
    ("出典",         "_source_link",      False),
]
DEC_COLUMNS = [
    ("削除",         "_delete_flag",      True),
    ("id",           "id",                False),
    ("内容",         "content",           True),
    ("決定日",       "decided_at",        True),
    ("確認済み",     "acknowledged",      True),
    ("出典",         "_source_link",      False),
]
RISK_COLUMNS = [
    ("id",           "id",                False),
    ("発生日",       "_extracted_date",   False),
    ("内容",         "content",           False),
    ("担当者",       "assignee",          False),
    ("期限",         "due_date",          False),
    ("MS",           "milestone_id",      False),
    ("対応状況",     "note",              False),
    ("出典",         "_source_link",      False),
]
ACH_COLUMNS = [
    ("アプリ",       "app",               False),
    ("実績",         "title",             False),
    ("分類",         "category",          False),
    ("時期",         "achieved_on",       False),
    ("確信度",       "confidence",        False),
    ("出典",         "evidence_ref",      False),
    ("根拠",         "evidence_quote",    False),
]
MS_COLUMNS = [
    ("milestone_id", False),
    ("名称",         False),
    ("期限",         False),
    ("状況",         False),
    ("完了/全体",    False),
    ("達成条件",     False),
]

HEADER_FILL_EDIT = PatternFill("solid", fgColor="D9E1F2")  # 編集可: 薄青
HEADER_FILL_RO   = PatternFill("solid", fgColor="E7E6E6")  # 読み取り: 薄灰
HEADER_FONT      = Font(bold=True)
RO_FILL          = PatternFill("solid", fgColor="F2F2F2")  # 読み取り列ハイライト
LINK_FONT        = Font(color="0563C1", underline="single")


# --------------------------------------------------------------------------- #
# 設定
# --------------------------------------------------------------------------- #
def load_report_config(config_path: Path) -> dict:
    if not config_path.exists():
        return {}
    with open(config_path, encoding="utf-8") as f:
        cfg = yaml.safe_load(f) or {}
    return cfg.get("report") or {}


def fetch_achievements(conn) -> list[dict]:
    """確定済み実績（status='confirmed'）を一覧取得する。"""
    rows = conn.execute(
        "SELECT app, title, category, achieved_on, confidence, evidence_ref, evidence_quote"
        " FROM achievements WHERE status='confirmed' AND COALESCE(deleted,0)=0"
        " ORDER BY app, achieved_on"
    ).fetchall()
    return [dict(r) for r in rows]


# --------------------------------------------------------------------------- #
# 出典 URL 解決
# --------------------------------------------------------------------------- #
def _build_meeting_url_map(rows: list[dict], minutes_dir: Path) -> dict[str, dict]:
    """meeting_id → {permalink, box_url} の辞書を返す。

    minutes/{kind}.db の instances.slack_file_permalink と upload_log.box_shared_url を引く。
    """
    kind_to_ids: dict[str, list[str]] = {}
    for r in rows:
        if r.get("source") != "meeting":
            continue
        kind = r.get("meeting_kind") or ""
        mid = r.get("meeting_id") or ""
        if kind and mid:
            kind_to_ids.setdefault(kind, []).append(mid)

    url_map: dict[str, dict] = {}
    for kind, meeting_ids in kind_to_ids.items():
        safe = re.sub(r"[^\w\-]", "_", kind)
        db_path = minutes_dir / f"{safe}.db"
        if not db_path.exists():
            continue
        try:
            conn_m = open_db(db_path)
            placeholders = ",".join("?" * len(meeting_ids))
            for r in conn_m.execute(
                f"SELECT meeting_id, slack_file_permalink FROM instances"
                f" WHERE meeting_id IN ({placeholders})",
                meeting_ids,
            ).fetchall():
                url_map.setdefault(r["meeting_id"], {})["permalink"] = r["slack_file_permalink"]
            # upload_log は存在しない DB もある
            try:
                for r in conn_m.execute(
                    f"SELECT meeting_id, box_shared_url FROM upload_log"
                    f" WHERE meeting_id IN ({placeholders})",
                    meeting_ids,
                ).fetchall():
                    url_map.setdefault(r["meeting_id"], {})["box_url"] = r["box_shared_url"]
            except Exception:
                pass
            conn_m.close()
        except Exception:
            pass
    return url_map


def _source_label(row: dict) -> str:
    if row.get("source") == "meeting":
        kind = row.get("meeting_kind") or ""
        held = row.get("meeting_held_at") or ""
        return f"{kind} ({held})" if held else (kind or "meeting")
    return "Slack"


def _source_url(row: dict, url_map: dict[str, dict]) -> str:
    if row.get("source") == "meeting":
        info = url_map.get(row.get("meeting_id") or "") or {}
        return info.get("box_url") or info.get("permalink") or ""
    ref = row.get("source_ref") or ""
    return ref if ref.startswith("http") else ""


def _extracted_date(row: dict) -> str:
    """会議由来は held_at、Slack 由来は extracted_at の日付部分のみを返す。"""
    raw = row.get("meeting_held_at") or row.get("extracted_at") or ""
    return raw[:10] if raw else ""


# --------------------------------------------------------------------------- #
# シート生成
# --------------------------------------------------------------------------- #
def _write_header(ws: Worksheet, columns, editable_cols: set[str]) -> None:
    for idx, col in enumerate(columns, start=1):
        name = col[0]
        ws.cell(row=1, column=idx, value=name).font = HEADER_FONT
        if name in editable_cols:
            ws.cell(row=1, column=idx).fill = HEADER_FILL_EDIT
        else:
            ws.cell(row=1, column=idx).fill = HEADER_FILL_RO
    ws.freeze_panes = "A2"
    ws.row_dimensions[1].height = 22


def _autosize(ws: Worksheet, columns, max_widths: dict[str, int]) -> None:
    for idx, col in enumerate(columns, start=1):
        name = col[0]
        cap = max_widths.get(name, 30)
        # 概算: ヘッダ+データの最大長（マルチバイトは 2 倍カウント）
        width = len(name) * 2
        for row in range(2, ws.max_row + 1):
            v = ws.cell(row=row, column=idx).value
            if v is None:
                continue
            s = str(v)
            w = sum(2 if ord(c) > 127 else 1 for c in s)
            if w > width:
                width = w
        ws.column_dimensions[get_column_letter(idx)].width = min(max(10, width + 2), cap)


def _add_dropdown(ws: Worksheet, columns, col_key: str, options: list[str],
                  last_row: int) -> None:
    """指定 key の列にドロップダウンリストを設定する。"""
    if last_row < 2:
        return
    col_idx = next((i for i, c in enumerate(columns, start=1) if c[1] == col_key), None)
    if col_idx is None:
        return
    letter = get_column_letter(col_idx)
    formula = '"' + ",".join(options) + '"'
    dv = DataValidation(type="list", formula1=formula, allow_blank=True,
                        showDropDown=False)  # False = ドロップダウン矢印を表示
    dv.add(f"{letter}2:{letter}{last_row}")
    ws.add_data_validation(dv)


def _set_source_cell(cell, item: dict, url_map: dict[str, dict]) -> None:
    """出典セルに表示ラベルを書き、URL があればハイパーリンクを設定する。"""
    label = _source_label(item)
    url = _source_url(item, url_map)
    cell.value = label
    if url:
        cell.hyperlink = url
        cell.font = LINK_FONT


def _set_cell_value(cell, key: str, item: dict, url_map: dict[str, dict]) -> None:
    """key に応じてセル値（または派生値）を設定する。"""
    if key == "_source_link":
        _set_source_cell(cell, item, url_map)
        return
    if key == "_extracted_date":
        cell.value = _extracted_date(item)
        return
    if key == "acknowledged":
        cell.value = "✓" if item.get("acknowledged_at") else ""
        return
    if key == "_delete_flag":
        # 削除済みは fetch から除外されているので常に空欄。
        # ユーザがここに 1/✓ 等を書き込むと sync 時に deleted=1 化される。
        cell.value = ""
        return
    cell.value = item.get(key)


def _build_action_items_sheet(ws: Worksheet, rows: list[dict],
                              url_map: dict[str, dict]) -> None:
    editable = {label for label, key, ed in AI_COLUMNS if ed}
    _write_header(ws, AI_COLUMNS, editable)
    for r_idx, item in enumerate(rows, start=2):
        for c_idx, (_label, key, ed) in enumerate(AI_COLUMNS, start=1):
            cell = ws.cell(row=r_idx, column=c_idx)
            _set_cell_value(cell, key, item, url_map)
            if not ed:
                cell.fill = RO_FILL
            cell.alignment = Alignment(wrap_text=True, vertical="top")
    last_row = len(rows) + 1
    _add_dropdown(ws, AI_COLUMNS, "_delete_flag", ["", "1"], last_row)
    _add_dropdown(ws, AI_COLUMNS, "status", ["open", "closed"], last_row)
    _autosize(ws, AI_COLUMNS, {"内容": 80, "対応状況": 60, "出典": 35})


def _build_decisions_sheet(ws: Worksheet, rows: list[dict],
                           url_map: dict[str, dict]) -> None:
    editable = {label for label, key, ed in DEC_COLUMNS if ed}
    _write_header(ws, DEC_COLUMNS, editable)
    for r_idx, item in enumerate(rows, start=2):
        for c_idx, (_label, key, ed) in enumerate(DEC_COLUMNS, start=1):
            cell = ws.cell(row=r_idx, column=c_idx)
            _set_cell_value(cell, key, item, url_map)
            if not ed:
                cell.fill = RO_FILL
            cell.alignment = Alignment(wrap_text=True, vertical="top")
    last_row = len(rows) + 1
    _add_dropdown(ws, DEC_COLUMNS, "_delete_flag", ["", "1"], last_row)
    _add_dropdown(ws, DEC_COLUMNS, "acknowledged", ["", "✓"], last_row)
    _autosize(ws, DEC_COLUMNS, {"内容": 80, "出典": 35})


def _build_achievements_sheet(ws: Worksheet, rows: list[dict]) -> None:
    editable = {label for label, key, ed in ACH_COLUMNS if ed}
    _write_header(ws, ACH_COLUMNS, editable)
    for r_idx, item in enumerate(rows, start=2):
        for c_idx, (_label, key, ed) in enumerate(ACH_COLUMNS, start=1):
            cell = ws.cell(row=r_idx, column=c_idx, value=item.get(key))
            if not ed:
                cell.fill = RO_FILL
            cell.alignment = Alignment(wrap_text=True, vertical="top")
    ws.protection.sheet = True  # 編集ブロック（ユーザー警告用）
    _autosize(ws, ACH_COLUMNS, {"実績": 60, "根拠": 80, "出典": 35})


def _build_risk_sheet(ws: Worksheet, rows: list[dict],
                      url_map: dict[str, dict]) -> None:
    _write_header(ws, RISK_COLUMNS, set())
    for r_idx, item in enumerate(rows, start=2):
        for c_idx, (_label, key, _ed) in enumerate(RISK_COLUMNS, start=1):
            cell = ws.cell(row=r_idx, column=c_idx)
            _set_cell_value(cell, key, item, url_map)
            cell.fill = RO_FILL
            cell.alignment = Alignment(wrap_text=True, vertical="top")
    ws.protection.sheet = True  # 編集ブロック（ユーザー警告用）
    _autosize(ws, RISK_COLUMNS, {"内容": 80, "対応状況": 60, "出典": 35})


def _milestone_status_label(m: dict, today: str) -> str:
    open_c = m["open_count"]
    closed_c = m["closed_count"]
    total = open_c + closed_c
    if m["status"] == "achieved":
        return "達成済"
    if not m["due_date"]:
        return "未着手" if total == 0 else "進行中"
    if m["due_date"] < today:
        return "遅延"
    if total == 0:
        return "未着手"
    pct = closed_c / total * 100 if total else 0
    if pct >= 80:
        return "進行中"
    if m["due_date"] <= date.today().replace(day=1).isoformat():
        return "要注意"
    return "進行中"


def _build_milestones_sheet(ws: Worksheet, milestones: list[dict],
                            today: str) -> None:
    _write_header(ws, [(c[0],) for c in MS_COLUMNS], set())
    for r_idx, m in enumerate(milestones, start=2):
        total = m["open_count"] + m["closed_count"]
        ratio = f"{m['closed_count']}/{total}" if total else "0/0"
        try:
            criteria = json.loads(m.get("success_criteria") or "[]")
            criteria_str = "\n".join(f"・{c}" for c in criteria)
        except Exception:
            criteria_str = ""
        values = [
            m["milestone_id"],
            m["name"],
            m["due_date"] or "",
            _milestone_status_label(m, today),
            ratio,
            criteria_str,
        ]
        for c_idx, v in enumerate(values, start=1):
            cell = ws.cell(row=r_idx, column=c_idx, value=v)
            cell.fill = RO_FILL
            cell.alignment = Alignment(wrap_text=True, vertical="top")
    ws.protection.sheet = True
    _autosize(ws, [(c[0],) for c in MS_COLUMNS], {"名称": 50, "達成条件": 80})


def build_workbook(
    action_items: list[dict],
    decisions: list[dict],
    risk_items: list[dict],
    milestones: list[dict],
    url_map: dict[str, dict],
    today: str,
    since: str | None,
    achievements: list[dict] | None = None,
) -> Workbook:
    wb = Workbook()
    # デフォルトシートを使い回し → 並び順: 現在地 / 要注意 / AI / 決定 / (実績)
    ws_ms = wb.active
    ws_ms.title = SHEET_MS
    ws_risk = wb.create_sheet(SHEET_RISK)
    ws_ai = wb.create_sheet(SHEET_AI)
    ws_dec = wb.create_sheet(SHEET_DEC)

    _build_milestones_sheet(ws_ms, milestones, today)
    _build_risk_sheet(ws_risk, risk_items, url_map)
    _build_action_items_sheet(ws_ai, action_items, url_map)
    _build_decisions_sheet(ws_dec, decisions, url_map)

    if achievements:
        ws_ach = wb.create_sheet(SHEET_ACH)
        _build_achievements_sheet(ws_ach, achievements)

    # ブックのプロパティ
    wb.properties.title = f"富岳NEXT 進捗レポート ({today})"
    wb.properties.creator = "pm_xlsx_report.py"
    wb.properties.description = (
        f"集計範囲: {since or '全期間'} / アクション {len(action_items)}件 / "
        f"決定 {len(decisions)}件 / 要注意 {len(risk_items)}件 / MS {len(milestones)}件"
        + (f" / 実績 {len(achievements)}件" if achievements else "")
    )
    return wb


# --------------------------------------------------------------------------- #
# Box CLI
# --------------------------------------------------------------------------- #
from box_cli import (
    box_get_or_create_shared_link,
    box_upload_or_version,
)


# --------------------------------------------------------------------------- #
# Canvas 目録
# --------------------------------------------------------------------------- #
def build_catalog_markdown(
    box_url: str,
    folder_id: str | None,
    filename: str,
    today: str,
    since: str | None,
    action_items: list[dict],
    decisions: list[dict],
    risk_items: list[dict],
    milestones: list[dict],
) -> str:
    since_note = f"{since} 以降" if since else "全期間"
    folder_url = f"https://app.box.com/folder/{folder_id}" if folder_id else ""

    ms_total = len(milestones)
    ms_achieved = sum(1 for m in milestones if m["status"] == "achieved")
    ms_overdue = sum(
        1 for m in milestones
        if m.get("due_date") and m["due_date"] < today and m["status"] != "achieved"
    )
    ms_active = ms_total - ms_achieved - ms_overdue

    overdue_ai = sum(
        1 for a in action_items
        if a.get("due_date") and a["due_date"] < today
    )
    no_due = sum(1 for a in action_items if not a.get("due_date"))

    ack_decisions = sum(1 for d in decisions if d.get("acknowledged_at"))

    lines = [
        f"# 富岳NEXT 進捗レポート目録 ({today})",
        "",
        f"📊 [最新版を開く ({filename})]({box_url})",
    ]
    if folder_url:
        lines.append(f"📁 [過去版フォルダ]({folder_url})")
    lines += [
        "",
        f"集計範囲: {since_note}（自動更新: {today}）",
        "",
        "## 現在地サマリ",
        f"- マイルストーン: 達成 **{ms_achieved}** / 進行中 **{ms_active}** / "
        f"遅延 **{ms_overdue}**（全 {ms_total} 件）",
        f"- 未完了アクションアイテム: **{len(action_items)}** 件"
        f"（期限超過: **{overdue_ai}** / 期限未設定: {no_due}）",
        f"- 要注意事項: **{len(risk_items)}** 件",
        f"- 直近の決定事項: **{len(decisions)}** 件（うち確認済み {ack_decisions}）",
        "",
        "## 編集方法",
        "- アクションアイテム / 決定事項 シートを直接編集してください",
        "- 次回 pm.db 更新時に自動で反映されます（要注意事項・現在地シートは読み取り専用）",
    ]
    return "\n".join(lines)


# --------------------------------------------------------------------------- #
# メイン
# --------------------------------------------------------------------------- #
def main() -> None:
    parser = argparse.ArgumentParser(
        description="pm.db → XLSX → Box アップロード → Canvas 目録投稿"
    )
    parser.add_argument("--db", default=str(DEFAULT_PM_DB),
                        help=f"pm.db のパス（デフォルト: {DEFAULT_PM_DB}）")
    parser.add_argument("--config", default=str(DEFAULT_CONFIG),
                        help=f"argus_config.yaml のパス（デフォルト: {DEFAULT_CONFIG}）")
    parser.add_argument("--canvas-id", default=None,
                        help="目録 Canvas ID（未指定時は report.canvas_id を参照）")
    parser.add_argument("--box-folder-id", default=None,
                        help="アップロード先 Box folder ID（未指定時は report.box_folder_id を参照）")
    parser.add_argument("--filename", default=None,
                        help=f"Box 上のファイル名（デフォルト: {DEFAULT_FILENAME}）")
    add_since_arg(parser)
    parser.add_argument("--show-acknowledged", action="store_true",
                        help="確認済み決定事項も含める（デフォルトは非表示）")
    parser.add_argument("--skip-upload", action="store_true",
                        help="Box アップロードをスキップ（ローカル生成のみ）")
    parser.add_argument("--skip-canvas", action="store_true",
                        help="Canvas 目録投稿をスキップ")
    parser.add_argument("--xlsx-out", default=None,
                        help="ローカル XLSX の保存先（指定しない場合は data/{filename}）")
    add_filter_arg(parser)
    add_dry_run_arg(parser)
    add_no_encrypt_arg(parser)
    args = parser.parse_args()

    db_path = Path(args.db)
    today = date.today().isoformat()

    log, close_log = make_logger(None)

    report_cfg = load_report_config(Path(args.config))
    canvas_id = args.canvas_id or resolve_report_canvas_id()
    folder_id = args.box_folder_id or report_cfg.get("box_folder_id")
    filename = args.filename or report_cfg.get("filename") or DEFAULT_FILENAME

    channel_ids, meeting_kinds = resolve_filter_presets(args.filter)

    log(f"[INFO] pm.db        : {db_path}")
    log(f"[INFO] since        : {args.since or '全期間'}")
    log(f"[INFO] box folder   : {folder_id or '(未設定)'}")
    log(f"[INFO] filename     : {filename}")
    log(f"[INFO] canvas       : {canvas_id or '(未設定)'}")

    conn = open_pm_db(db_path, no_encrypt=args.no_encrypt)
    action_items = fetch_open_action_items(conn, args.since, channel_ids, meeting_kinds)
    decisions = fetch_recent_decisions(conn, args.since,
                                       show_acknowledged=args.show_acknowledged,
                                       channel_ids=channel_ids,
                                       meeting_kinds=meeting_kinds)
    risk_items = detect_risk_items(action_items)
    milestones = fetch_milestone_progress(conn)
    achievements = fetch_achievements(conn)
    conn.close()

    minutes_dir = db_path.parent / "minutes"
    url_map = _build_meeting_url_map(action_items + decisions, minutes_dir)

    log(f"[INFO] アクション   : {len(action_items)}件 (要注意 {len(risk_items)}件)")
    log(f"[INFO] 決定事項     : {len(decisions)}件")
    log(f"[INFO] マイルストーン: {len(milestones)}件")
    log(f"[INFO] 実績         : {len(achievements)}件")

    wb = build_workbook(action_items, decisions, risk_items, milestones,
                        url_map, today, args.since, achievements=achievements)

    out_path = Path(args.xlsx_out) if args.xlsx_out else REPO_ROOT / "data" / filename
    out_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(out_path)
    log(f"[INFO] 生成: {out_path}")

    if args.dry_run:
        log("[INFO] dry-run のため Box アップロード・Canvas 投稿をスキップ")
        close_log()
        return

    box_url = ""
    if not args.skip_upload:
        if not folder_id:
            log("[WARN] box_folder_id が未設定のため Box アップロードをスキップ")
        else:
            try:
                file_id = box_upload_or_version(out_path, folder_id, filename, log)
                box_url = box_get_or_create_shared_link(file_id, log)
                log(f"[OK]   Box file_id={file_id} url={box_url}")
            except (subprocess.CalledProcessError, subprocess.TimeoutExpired,
                    RuntimeError) as e:
                log(f"[ERROR] Box アップロード失敗: {e}")

    if not args.skip_canvas:
        if not canvas_id:
            log("[WARN] canvas_id が未設定のため Canvas 投稿をスキップ")
        elif not box_url:
            log("[WARN] Box URL が取得できなかったため Canvas 投稿をスキップ")
        else:
            md = build_catalog_markdown(
                box_url, folder_id, filename, today, args.since,
                action_items, decisions, risk_items, milestones,
            )
            md = sanitize_for_canvas(md)
            log("\n" + "=" * 60)
            log(md)
            log("=" * 60)
            post_to_canvas(canvas_id, md)

    close_log()


if __name__ == "__main__":
    main()
