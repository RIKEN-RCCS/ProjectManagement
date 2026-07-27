#!/usr/bin/env python3
"""
pm_xlsx_sync.py

Box にアップロードされた進捗レポート XLSX をダウンロードし、
アクションアイテム / 決定事項 シートの編集内容を pm.db に反映する。
要注意事項 / プロジェクトの現在地 シートは読み取り専用として無視する。

Usage:
    python3 scripts/pm_xlsx_sync.py
    python3 scripts/pm_xlsx_sync.py --dry-run
    python3 scripts/pm_xlsx_sync.py --xlsx data/pm_report.xlsx
"""
from __future__ import annotations

import argparse
import subprocess
import sys
import tempfile
from collections import defaultdict
from datetime import UTC, datetime
from pathlib import Path

import yaml
from openpyxl import load_workbook

sys.path.insert(0, str(Path(__file__).resolve().parent.parent))
from cli_utils import add_dry_run_arg, add_no_encrypt_arg, make_logger
from db_utils import open_pm_db
from pm_relink import write_audit_log

REPO_ROOT = Path(__file__).resolve().parent.parent.parent
DEFAULT_PM_DB = REPO_ROOT / "data" / "pm.db"
DEFAULT_CONFIG = REPO_ROOT / "data" / "argus_config.yaml"
DEFAULT_FILENAME = "pm_report.xlsx"

SHEET_AI = "アクションアイテム"
SHEET_DEC = "決定事項"

# (シート列名, DB 列名) — pm_xlsx_report.py の AI_COLUMNS / DEC_COLUMNS と整合
AI_FIELD_MAP = {
    "内容":      "content",
    "担当者":    "assignee",
    "期限":      "due_date",
    "MS":        "milestone_id",
    "状況":      "status",
    "対応状況":  "note",
    "削除":      "deleted",
}
AI_NULLABLE = {"assignee", "due_date", "milestone_id", "note"}

DEC_FIELD_MAP = {
    "内容":      "content",
    "決定日":    "decided_at",
    "確認済み":  "acknowledged",   # 特殊扱い: ✓/x/y/true → acknowledged_at
    "削除":      "deleted",
}

# 削除フラグとして 1 扱いになる文字列
_DELETE_TRUE_TOKENS = {"✓", "x", "y", "yes", "true", "1", "○", "済", "削除"}


def _parse_delete_flag(raw) -> int:
    if raw is None:
        return 0
    s = str(raw).strip().lower()
    if not s:
        return 0
    return 1 if s in _DELETE_TRUE_TOKENS else 0


# --------------------------------------------------------------------------- #
# Box CLI
# --------------------------------------------------------------------------- #
from box_cli import box_download, box_find_file, box_get_file_modified_at


def fetch_xlsx_from_box(folder_id: str, filename: str, dest_dir: Path,
                       log) -> tuple[Path | None, str | None]:
    """Box の最新ファイルをダウンロード。存在しなければ (None, None) を返す。

    戻り値の第2要素は Box 上の modified_at（ISO8601、鮮度ガードのフォールバック用）。
    取得失敗時は None。
    """
    try:
        file_id = box_find_file(folder_id, filename)
    except (subprocess.CalledProcessError, subprocess.TimeoutExpired) as e:
        log(f"[WARN] box folders:items 失敗 (folder_id={folder_id}): {e}")
        return None, None
    if not file_id:
        log(f"[INFO] Box フォルダに {filename} が存在しません（初回 or 未生成）")
        return None, None
    dest = dest_dir / filename
    try:
        box_download(file_id, dest, log)
    except (subprocess.CalledProcessError, subprocess.TimeoutExpired) as e:
        log(f"[WARN] box files:download 失敗: {e}")
        return None, None
    return dest, box_get_file_modified_at(file_id, log)


# --------------------------------------------------------------------------- #
# 鮮度ガード: シートより新しい pm.db 側の変更を巻き戻さない
# --------------------------------------------------------------------------- #
def _ensure_aware_utc(dt: datetime) -> datetime:
    """naive datetime は UTC とみなして aware 化し、UTC に変換する。"""
    if dt.tzinfo is None:
        dt = dt.replace(tzinfo=UTC)
    return dt.astimezone(UTC)


def _parse_iso_dt(s: str | None) -> datetime | None:
    """ISO8601 文字列を aware UTC datetime に変換する。パース不能なら None。"""
    if not s:
        return None
    try:
        dt = datetime.fromisoformat(str(s).replace("Z", "+00:00"))
    except ValueError:
        return None
    return _ensure_aware_utc(dt)


def _fetch_latest_audit(
    conn, table_name: str, record_ids: list[int],
) -> dict[tuple[int, str], tuple[datetime | None, str, str]]:
    """(record_id, field) → (パース済み changed_at, changed_at 原文, source) の最新変更を返す。

    source='xlsx_sync'（自分自身の過去同期）は「新しい変更」とみなさないため除外する。
    文字列の ORDER BY には依存せず、Python 側で全件パースして最大値を求める。
    パース不能な changed_at が候補に一つでもあれば、その (record_id, field) は
    安全側として「パース不能」（先頭要素 None）を保持する。
    """
    if not record_ids:
        return {}
    placeholders = ",".join("?" * len(record_ids))
    rows = conn.execute(
        f"SELECT record_id, field, changed_at, source FROM audit_log"
        f" WHERE table_name = ? AND record_id IN ({placeholders})"
        f" AND (source IS NULL OR source != 'xlsx_sync')",
        [table_name, *[str(r) for r in record_ids]],
    ).fetchall()
    latest: dict[tuple[int, str], tuple[datetime | None, str, str]] = {}
    for r in rows:
        key = (int(r["record_id"]), r["field"])
        dt = _parse_iso_dt(r["changed_at"])
        existing = latest.get(key)
        if existing is None:
            latest[key] = (dt, r["changed_at"], r["source"])
        elif dt is None:
            # パース不能な候補が一つでもあれば安全側を維持する
            latest[key] = (None, r["changed_at"], r["source"])
        elif existing[0] is not None and dt > existing[0]:
            latest[key] = (dt, r["changed_at"], r["source"])
    return latest


def _guard_blocked(
    latest_audit: dict[tuple[int, str], tuple[datetime | None, str, str]],
    rid: int, field: str, base_time: datetime, label: str, log,
) -> bool:
    """(rid, field) が鮮度ガードでスキップ対象かどうかを判定する（True ならスキップ）。"""
    key = (rid, field)
    if key not in latest_audit:
        return False
    dt, changed_at, source = latest_audit[key]
    if dt is None:
        log(f"[WARN] {label} #{rid} の {field} 編集は audit_log.changed_at={changed_at!r} を"
            f"パースできないため安全側で破棄しました。意図的に上書きする場合のみ --force を指定してください")
        return True
    if dt > base_time:
        log(f"[WARN] {label} #{rid} のシート編集（{field}）は、pm.db 側の新しい変更"
            f"（source={source}, changed_at={changed_at}）により破棄されました。"
            f"意図的に古いシートで上書きする場合のみ --force を指定してください"
            f"（patrol の自動クローズ結果を再度巻き戻す点に注意）")
        return True
    return False


# --------------------------------------------------------------------------- #
# XLSX パース
# --------------------------------------------------------------------------- #
def _normalize_value(v) -> str | None:
    """セル値を文字列化。None / 空白 → None。"""
    if v is None:
        return None
    if isinstance(v, datetime):
        return v.date().isoformat()
    s = str(v).strip()
    return s if s else None


def _read_sheet(wb, sheet_name: str, field_map: dict[str, str]) -> dict[int, dict]:
    """シートを読み、id → {db_field: value} を返す。

    id が空の行はスキップ（新規行は未対応 — 既存編集のみ反映）。
    """
    if sheet_name not in wb.sheetnames:
        return {}
    ws = wb[sheet_name]
    header = [(c.value or "").strip() if c.value else "" for c in ws[1]]
    if "id" not in header:
        return {}
    id_col = header.index("id")

    # シート列名 → field_map の DB 列名を引いて、列番号と DB 列名を対応付ける
    col_to_db: list[tuple[int, str]] = []
    for idx, label in enumerate(header):
        db_field = field_map.get(label)
        if db_field:
            col_to_db.append((idx, db_field))

    rows: dict[int, dict] = {}
    for row in ws.iter_rows(min_row=2, values_only=True):
        if id_col >= len(row):
            continue
        raw_id = row[id_col]
        if raw_id is None or str(raw_id).strip() == "":
            continue
        try:
            rid = int(raw_id)
        except (TypeError, ValueError):
            continue
        rec: dict = {}
        for idx, db_field in col_to_db:
            if idx >= len(row):
                continue
            val = _normalize_value(row[idx])
            rec[db_field] = val
        rows[rid] = rec
    return rows


# --------------------------------------------------------------------------- #
# pm.db 反映
# --------------------------------------------------------------------------- #
def _coerce_for_field(table: str, field: str, val):
    """空文字を NULL に揃える。型変換が必要なフィールドはここで対応。"""
    if val is None:
        return None
    if val == "":
        return None
    return val


def _apply_action_items(
    conn, rows: dict[int, dict], dry_run: bool, log, base_time: datetime | None = None,
) -> tuple[int, int, int]:
    if not rows:
        return 0, 0, 0
    fields = list(set(f for r in rows.values() for f in r.keys()))
    if not fields:
        return 0, 0, 0
    placeholders = ",".join("?" * len(rows))
    field_list = ", ".join(fields)
    current = {
        r["id"]: dict(r) for r in conn.execute(
            f"SELECT id, {field_list} FROM action_items WHERE id IN ({placeholders})",
            list(rows.keys()),
        ).fetchall()
    }
    latest_audit = _fetch_latest_audit(conn, "action_items", list(rows.keys())) if base_time is not None else {}
    changes: list[tuple[int, str, object, object]] = []
    guard_skipped_rows: set[int] = set()
    for rid, new_vals in rows.items():
        if rid not in current:
            log(f"[WARN] action_items id={rid} は DB に存在しません。スキップ")
            continue
        cur = current[rid]
        for field, new_val in new_vals.items():
            if field == "deleted":
                # XLSX の「削除」列が空/0/Falseの場合は上書きしない。
                # ユーザーが明示的に削除フラグを立てた場合（✓/x/1/削除等）のみ
                # pm.db の deleted 列を更新する。これにより、Argus Console 等で
                # 既に deleted=1 になったレコードが XLSX 同期で復活するのを防ぐ。
                if not new_val or str(new_val).strip() in ("", "0", "False", "false"):
                    continue
                new_flag = _parse_delete_flag(new_val)
                old_val = cur.get("deleted") or 0
                if old_val == new_flag:
                    continue
                if base_time is not None and _guard_blocked(latest_audit, rid, "deleted", base_time, "AI", log):
                    guard_skipped_rows.add(rid)
                    continue
                changes.append((rid, "deleted", old_val, new_flag))
                continue
            coerced_val = _coerce_for_field("action_items", field, new_val)
            if field in AI_NULLABLE and (coerced_val == "" or coerced_val is None):
                coerced_val = None
            old_val = cur.get(field)
            # status は open/closed のみ受け付ける
            if field == "status" and coerced_val and coerced_val not in ("open", "closed"):
                continue
            if (old_val or None) == (coerced_val or None):
                continue
            if base_time is not None and _guard_blocked(latest_audit, rid, field, base_time, "AI", log):
                guard_skipped_rows.add(rid)
                continue
            changes.append((rid, field, old_val, coerced_val))
    if not changes:
        return 0, 0, len(guard_skipped_rows)
    by_item: dict[int, list] = defaultdict(list)
    for rid, f, ov, nv in changes:
        by_item[rid].append((f, ov, nv))
    for rid in sorted(by_item):
        log(f"  [AI ] id={rid}")
        for f, ov, nv in by_item[rid]:
            log(f"    {f:<14}: {ov!s:<30} → {nv!s}")
    if not dry_run:
        for rid, f, ov, nv in changes:
            write_audit_log(conn, "action_items", rid, f, ov, nv, "xlsx_sync")
        for rid, fcs in by_item.items():
            set_clause = ", ".join(f"{f} = ?" for f, _, _ in fcs)
            values = [nv for _, _, nv in fcs] + [rid]
            conn.execute(f"UPDATE action_items SET {set_clause} WHERE id = ?", values)
    return len(changes), len(by_item), len(guard_skipped_rows)


def _apply_decisions(
    conn, rows: dict[int, dict], dry_run: bool, log, base_time: datetime | None = None,
) -> tuple[int, int, int]:
    if not rows:
        return 0, 0, 0
    # acknowledged は特殊扱い: ✓/y/yes/true → acknowledged_at = today、空 → NULL
    today = datetime.now(UTC).date().isoformat()
    by_item: dict[int, list] = defaultdict(list)
    guard_skipped_rows: set[int] = set()
    db_fields = ["content", "decided_at", "acknowledged_at", "deleted"]
    placeholders = ",".join("?" * len(rows))
    field_list = ", ".join(db_fields)
    current = {
        r["id"]: dict(r) for r in conn.execute(
            f"SELECT id, {field_list} FROM decisions WHERE id IN ({placeholders})",
            list(rows.keys()),
        ).fetchall()
    }
    latest_audit = _fetch_latest_audit(conn, "decisions", list(rows.keys())) if base_time is not None else {}
    for rid, new_vals in rows.items():
        if rid not in current:
            log(f"[WARN] decisions id={rid} は DB に存在しません。スキップ")
            continue
        cur = current[rid]
        for label, new_val in new_vals.items():
            if label == "acknowledged":
                # XLSX の「確認済み」列が空の場合は上書きしない。
                # deleted と同様、Web で一括確認した acknowledged_at が
                # XLSX 同期でリセットされるのを防ぐ。
                if not new_val or str(new_val).strip() in ("", "0", "False", "false"):
                    continue
                ack = bool(new_val) and str(new_val).strip().lower() in ("✓", "x", "y", "yes", "true", "1", "○", "済")
                old_val = cur.get("acknowledged_at")
                new_db = today if ack else None
                # 既に acknowledged 済みなら新規日付で上書きしない
                if ack and old_val:
                    continue
                if (old_val or None) == (new_db or None):
                    continue
                if base_time is not None and _guard_blocked(
                    latest_audit, rid, "acknowledged_at", base_time, "DEC", log
                ):
                    guard_skipped_rows.add(rid)
                    continue
                by_item[rid].append(("acknowledged_at", old_val, new_db))
            elif label == "deleted":
                # XLSX の「削除」列が空/0/Falseの場合は上書きしない（action_items側と同様）。
                if not new_val or str(new_val).strip() in ("", "0", "False", "false"):
                    continue
                new_flag = _parse_delete_flag(new_val)
                old_flag = cur.get("deleted") or 0
                if old_flag == new_flag:
                    continue
                if base_time is not None and _guard_blocked(latest_audit, rid, "deleted", base_time, "DEC", log):
                    guard_skipped_rows.add(rid)
                    continue
                by_item[rid].append(("deleted", old_flag, new_flag))
            else:
                db_field = {"内容": "content", "決定日": "decided_at"}.get(label)
                # 既に DEC_FIELD_MAP で db_field 化されている想定
                if label in ("content", "decided_at"):
                    db_field = label
                if not db_field:
                    continue
                old_val = cur.get(db_field)
                if (old_val or None) == (new_val or None):
                    continue
                if base_time is not None and _guard_blocked(latest_audit, rid, db_field, base_time, "DEC", log):
                    guard_skipped_rows.add(rid)
                    continue
                by_item[rid].append((db_field, old_val, new_val))
    if not by_item:
        return 0, 0, len(guard_skipped_rows)
    total_changes = sum(len(v) for v in by_item.values())
    for rid in sorted(by_item):
        log(f"  [DEC] id={rid}")
        for f, ov, nv in by_item[rid]:
            log(f"    {f:<16}: {ov!s:<30} → {nv!s}")
    if not dry_run:
        for rid, fcs in by_item.items():
            for f, ov, nv in fcs:
                write_audit_log(conn, "decisions", rid, f, ov, nv, "xlsx_sync")
            set_clause = ", ".join(f"{f} = ?" for f, _, _ in fcs)
            values = [nv for _, _, nv in fcs] + [rid]
            conn.execute(f"UPDATE decisions SET {set_clause} WHERE id = ?", values)
    return total_changes, len(by_item), len(guard_skipped_rows)


# --------------------------------------------------------------------------- #
# 設定
# --------------------------------------------------------------------------- #
def load_report_config(config_path: Path) -> dict:
    if not config_path.exists():
        return {}
    with open(config_path, encoding="utf-8") as f:
        cfg = yaml.safe_load(f) or {}
    return cfg.get("report") or {}


# --------------------------------------------------------------------------- #
# メイン
# --------------------------------------------------------------------------- #
def main() -> None:
    parser = argparse.ArgumentParser(
        description="Box の進捗レポート XLSX を pm.db に反映する"
    )
    parser.add_argument("--db", default=str(DEFAULT_PM_DB),
                        help=f"pm.db のパス（デフォルト: {DEFAULT_PM_DB}）")
    parser.add_argument("--config", default=str(DEFAULT_CONFIG))
    parser.add_argument("--xlsx", default=None,
                        help="ローカル XLSX を直接読む（指定時は Box ダウンロードをスキップ）")
    parser.add_argument("--box-folder-id", default=None,
                        help="Box folder ID（未指定時は report.box_folder_id）")
    parser.add_argument("--filename", default=None,
                        help=f"Box 上のファイル名（デフォルト: {DEFAULT_FILENAME}）")
    parser.add_argument("--force", action="store_true",
                        help="鮮度ガードを無視して従来どおり全行同期する")
    add_dry_run_arg(parser)
    add_no_encrypt_arg(parser)
    args = parser.parse_args()

    log, _ = make_logger(None)
    db_path = Path(args.db)
    report_cfg = load_report_config(Path(args.config))
    folder_id = args.box_folder_id or report_cfg.get("box_folder_id")
    filename = args.filename or report_cfg.get("filename") or DEFAULT_FILENAME

    fallback_ts: datetime | None
    if args.xlsx:
        xlsx_path = Path(args.xlsx)
        if not xlsx_path.exists():
            log(f"[ERROR] ファイルが見つかりません: {xlsx_path}")
            sys.exit(1)
        fallback_ts = datetime.fromtimestamp(xlsx_path.stat().st_mtime, tz=UTC)
        fallback_label = "ローカルファイル mtime"
    else:
        if not folder_id:
            log("[INFO] box_folder_id が未設定のため XLSX 同期をスキップ")
            return
        with tempfile.TemporaryDirectory() as td:
            xlsx_path, modified_at_raw = fetch_xlsx_from_box(folder_id, filename, Path(td), log)
            if xlsx_path is None:
                return
            fallback_ts = _parse_iso_dt(modified_at_raw)
            fallback_label = "Box modified_at"
            return _process_xlsx(xlsx_path, db_path, args, log, fallback_ts, fallback_label)

    _process_xlsx(xlsx_path, db_path, args, log, fallback_ts, fallback_label)


def _process_xlsx(
    xlsx_path: Path, db_path: Path, args, log,
    fallback_ts: datetime | None = None, fallback_label: str = "",
) -> None:
    log(f"[INFO] XLSX        : {xlsx_path}")
    log(f"[INFO] pm.db       : {db_path}")

    # 議事録 DB (data/minutes/*.db) は pm.db とスキーマが異なるため対象外
    if "minutes" in str(db_path):
        log(f"[ERROR] --db に議事録 DB が指定されています: {db_path}")
        log("[ERROR] pm_xlsx_sync.py は pm.db 専用です。--db data/pm.db を指定してください。")
        sys.exit(1)

    wb = load_workbook(xlsx_path, data_only=True)

    # 鮮度ガードの基準時刻: 主基準はワークブックの作成打刻（openpyxl が
    # 保存時に dcterms:created として書き込む naive UTC。pm_xlsx_report.py が
    # エクスポート毎に Workbook() を新規生成するため「pm.db からエクスポート
    # した時点」を表し、Excel 手編集後も core.xml の created は保持される）。
    # Box modified_at → ローカル mtime は取得失敗時のフォールバック。
    sheet_ts: datetime | None
    basis: str | None
    created = wb.properties.created
    if created is not None:
        sheet_ts = _ensure_aware_utc(created)
        basis = "ワークブック作成日時 (dcterms:created)"
    elif fallback_ts is not None:
        sheet_ts = fallback_ts
        basis = fallback_label or "フォールバック"
    else:
        sheet_ts = None
        basis = None

    if args.force:
        log("[INFO] --force 指定のため鮮度ガードを無視して同期します")
        guard_base_time = None
    elif sheet_ts is None:
        log("[WARN] シート基準時刻を取得できないため鮮度ガード無効で実行します")
        guard_base_time = None
    else:
        log(f"[INFO] シート基準時刻: {sheet_ts.isoformat()}（基準: {basis}）")
        guard_base_time = sheet_ts

    ai_rows = _read_sheet(wb, SHEET_AI, AI_FIELD_MAP)
    dec_rows = _read_sheet(wb, SHEET_DEC, DEC_FIELD_MAP)
    log(f"[INFO] アクション編集候補: {len(ai_rows)}件 / 決定事項編集候補: {len(dec_rows)}件")

    conn = open_pm_db(db_path, no_encrypt=args.no_encrypt)
    try:
        ai_changes, ai_items, ai_skipped = _apply_action_items(
            conn, ai_rows, args.dry_run, log, guard_base_time)
        dec_changes, dec_items, dec_skipped = _apply_decisions(
            conn, dec_rows, args.dry_run, log, guard_base_time)
        if not args.dry_run:
            conn.commit()
    finally:
        conn.close()

    log(f"[INFO] action_items: {ai_changes} 変更 / {ai_items} 件")
    log(f"[INFO] decisions    : {dec_changes} 変更 / {dec_items} 件")
    total_skipped = ai_skipped + dec_skipped
    if total_skipped:
        log(f"[WARN] {total_skipped} 行をシートより新しい変更のためスキップ")
    if args.dry_run:
        log("[INFO] dry-run のため DB は変更していません")


if __name__ == "__main__":
    main()
