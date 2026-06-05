#!/usr/bin/env python3
"""
pm_xlsx_sync.py

Box にアップロードされた進捗レポート XLSX をダウンロードし、
アクションアイテム / 決定事項 シートの編集内容を pm.db に反映する。
要注意事項 / プロジェクトの現在地 シートは読み取り専用として無視する。

Usage:
    python3 scripts/pm_xlsx_sync.py
    python3 scripts/pm_xlsx_sync.py --dry-run
    python3 scripts/pm_xlsx_sync.py --xlsx data/pm_report.xlsx
"""
from __future__ import annotations

import argparse
import json
import subprocess
import sys
import tempfile
from collections import defaultdict
from datetime import datetime, timezone
from pathlib import Path

import yaml
from openpyxl import load_workbook

sys.path.insert(0, str(Path(__file__).resolve().parent))
from db_utils import open_pm_db
from cli_utils import add_dry_run_arg, add_no_encrypt_arg, make_logger
from pm_relink import write_audit_log

REPO_ROOT = Path(__file__).resolve().parent.parent
DEFAULT_PM_DB = REPO_ROOT / "data" / "pm.db"
DEFAULT_CONFIG = REPO_ROOT / "data" / "argus_config.yaml"
DEFAULT_FILENAME = "pm_report.xlsx"

SHEET_AI = "アクションアイテム"
SHEET_DEC = "決定事項"

# (シート列名, DB 列名) — pm_xlsx_report.py の AI_COLUMNS / DEC_COLUMNS と整合
AI_FIELD_MAP = {
    "内容":      "content",
    "担当者":    "assignee",
    "期限":      "due_date",
    "MS":        "milestone_id",
    "状況":      "status",
    "対応状況":  "note",
    "削除":      "deleted",
}
AI_NULLABLE = {"assignee", "due_date", "milestone_id", "note"}

DEC_FIELD_MAP = {
    "内容":      "content",
    "決定日":    "decided_at",
    "確認済み":  "acknowledged",   # 特殊扱い: ✓/x/y/true → acknowledged_at
    "削除":      "deleted",
}

# 削除フラグとして 1 扱いになる文字列
_DELETE_TRUE_TOKENS = {"✓", "x", "y", "yes", "true", "1", "○", "済", "削除"}


def _parse_delete_flag(raw) -> int:
    if raw is None:
        return 0
    s = str(raw).strip().lower()
    if not s:
        return 0
    return 1 if s in _DELETE_TRUE_TOKENS else 0


# --------------------------------------------------------------------------- #
# Box CLI
# --------------------------------------------------------------------------- #
def _box_json(cmd: list[str], timeout: int = 120):
    raw = subprocess.check_output(cmd, text=True, timeout=timeout)
    return json.loads(raw)


def box_find_file(folder_id: str, filename: str) -> str | None:
    items = _box_json(
        ["box", "folders:items", folder_id, "--json", "--fields", "name,type"],
        timeout=60,
    )
    for item in items:
        if item.get("type") == "file" and item.get("name") == filename:
            return str(item.get("id"))
    return None


def box_download(file_id: str, dest: Path, log) -> None:
    log(f"  [BOX] ダウンロード: file_id={file_id} → {dest}")
    subprocess.check_call(
        ["box", "files:download", file_id, "--destination", str(dest.parent),
         "--save-as", dest.name, "--overwrite"],
        timeout=300,
    )


def fetch_xlsx_from_box(folder_id: str, filename: str, dest_dir: Path,
                       log) -> Path | None:
    """Box の最新ファイルをダウンロード。存在しなければ None を返す。"""
    try:
        file_id = box_find_file(folder_id, filename)
    except (subprocess.CalledProcessError, subprocess.TimeoutExpired) as e:
        log(f"[WARN] box folders:items 失敗 (folder_id={folder_id}): {e}")
        return None
    if not file_id:
        log(f"[INFO] Box フォルダに {filename} が存在しません（初回 or 未生成）")
        return None
    dest = dest_dir / filename
    try:
        box_download(file_id, dest, log)
    except (subprocess.CalledProcessError, subprocess.TimeoutExpired) as e:
        log(f"[WARN] box files:download 失敗: {e}")
        return None
    return dest


# --------------------------------------------------------------------------- #
# XLSX パース
# --------------------------------------------------------------------------- #
def _normalize_value(v) -> str | None:
    """セル値を文字列化。None / 空白 → None。"""
    if v is None:
        return None
    if isinstance(v, datetime):
        return v.date().isoformat()
    s = str(v).strip()
    return s if s else None


def _read_sheet(wb, sheet_name: str, field_map: dict[str, str]) -> dict[int, dict]:
    """シートを読み、id → {db_field: value} を返す。

    id が空の行はスキップ（新規行は未対応 — 既存編集のみ反映）。
    """
    if sheet_name not in wb.sheetnames:
        return {}
    ws = wb[sheet_name]
    header = [(c.value or "").strip() if c.value else "" for c in ws[1]]
    if "id" not in header:
        return {}
    id_col = header.index("id")

    # シート列名 → field_map の DB 列名を引いて、列番号と DB 列名を対応付ける
    col_to_db: list[tuple[int, str]] = []
    for idx, label in enumerate(header):
        db_field = field_map.get(label)
        if db_field:
            col_to_db.append((idx, db_field))

    rows: dict[int, dict] = {}
    for row in ws.iter_rows(min_row=2, values_only=True):
        if id_col >= len(row):
            continue
        raw_id = row[id_col]
        if raw_id is None or str(raw_id).strip() == "":
            continue
        try:
            rid = int(raw_id)
        except (TypeError, ValueError):
            continue
        rec: dict = {}
        for idx, db_field in col_to_db:
            if idx >= len(row):
                continue
            val = _normalize_value(row[idx])
            rec[db_field] = val
        rows[rid] = rec
    return rows


# --------------------------------------------------------------------------- #
# pm.db 反映
# --------------------------------------------------------------------------- #
def _coerce_for_field(table: str, field: str, val):
    """空文字を NULL に揃える。型変換が必要なフィールドはここで対応。"""
    if val is None:
        return None
    if val == "":
        return None
    return val


def _apply_action_items(conn, rows: dict[int, dict], dry_run: bool, log) -> tuple[int, int]:
    if not rows:
        return 0, 0
    fields = list(set(f for r in rows.values() for f in r.keys()))
    if not fields:
        return 0, 0
    placeholders = ",".join("?" * len(rows))
    field_list = ", ".join(fields)
    current = {
        r["id"]: dict(r) for r in conn.execute(
            f"SELECT id, {field_list} FROM action_items WHERE id IN ({placeholders})",
            list(rows.keys()),
        ).fetchall()
    }
    changes: list[tuple[int, str, object, object]] = []
    for rid, new_vals in rows.items():
        if rid not in current:
            log(f"[WARN] action_items id={rid} は DB に存在しません。スキップ")
            continue
        cur = current[rid]
        for field, new_val in new_vals.items():
            if field == "deleted":
                new_val = _parse_delete_flag(new_val)
                old_val = cur.get("deleted") or 0
                if old_val != new_val:
                    changes.append((rid, "deleted", old_val, new_val))
                continue
            new_val = _coerce_for_field("action_items", field, new_val)
            if field in AI_NULLABLE and (new_val == "" or new_val is None):
                new_val = None
            old_val = cur.get(field)
            # status は open/closed のみ受け付ける
            if field == "status" and new_val and new_val not in ("open", "closed"):
                continue
            if (old_val or None) != (new_val or None):
                changes.append((rid, field, old_val, new_val))
    if not changes:
        return 0, 0
    by_item: dict[int, list] = defaultdict(list)
    for rid, f, ov, nv in changes:
        by_item[rid].append((f, ov, nv))
    for rid in sorted(by_item):
        log(f"  [AI ] id={rid}")
        for f, ov, nv in by_item[rid]:
            log(f"    {f:<14}: {ov!s:<30} → {nv!s}")
    if not dry_run:
        for rid, f, ov, nv in changes:
            write_audit_log(conn, "action_items", rid, f, ov, nv, "xlsx_sync")
        for rid, fcs in by_item.items():
            set_clause = ", ".join(f"{f} = ?" for f, _, _ in fcs)
            values = [nv for _, _, nv in fcs] + [rid]
            conn.execute(f"UPDATE action_items SET {set_clause} WHERE id = ?", values)
    return len(changes), len(by_item)


def _apply_decisions(conn, rows: dict[int, dict], dry_run: bool, log) -> tuple[int, int]:
    if not rows:
        return 0, 0
    # acknowledged は特殊扱い: ✓/y/yes/true → acknowledged_at = today、空 → NULL
    today = datetime.now(timezone.utc).date().isoformat()
    by_item: dict[int, list] = defaultdict(list)
    db_fields = ["content", "decided_at", "acknowledged_at", "deleted"]
    placeholders = ",".join("?" * len(rows))
    field_list = ", ".join(db_fields)
    current = {
        r["id"]: dict(r) for r in conn.execute(
            f"SELECT id, {field_list} FROM decisions WHERE id IN ({placeholders})",
            list(rows.keys()),
        ).fetchall()
    }
    for rid, new_vals in rows.items():
        if rid not in current:
            log(f"[WARN] decisions id={rid} は DB に存在しません。スキップ")
            continue
        cur = current[rid]
        for label, new_val in new_vals.items():
            if label == "acknowledged":
                ack = bool(new_val) and str(new_val).strip().lower() in ("✓", "x", "y", "yes", "true", "1", "○", "済")
                old_val = cur.get("acknowledged_at")
                new_db = today if ack else None
                # 既に acknowledged 済みなら新規日付で上書きしない
                if ack and old_val:
                    continue
                if (old_val or None) != (new_db or None):
                    by_item[rid].append(("acknowledged_at", old_val, new_db))
            elif label == "deleted":
                new_flag = _parse_delete_flag(new_val)
                old_flag = cur.get("deleted") or 0
                if old_flag != new_flag:
                    by_item[rid].append(("deleted", old_flag, new_flag))
            else:
                db_field = {"内容": "content", "決定日": "decided_at"}.get(label)
                # 既に DEC_FIELD_MAP で db_field 化されている想定
                if label in ("content", "decided_at"):
                    db_field = label
                if not db_field:
                    continue
                old_val = cur.get(db_field)
                if (old_val or None) != (new_val or None):
                    by_item[rid].append((db_field, old_val, new_val))
    if not by_item:
        return 0, 0
    total_changes = sum(len(v) for v in by_item.values())
    for rid in sorted(by_item):
        log(f"  [DEC] id={rid}")
        for f, ov, nv in by_item[rid]:
            log(f"    {f:<16}: {ov!s:<30} → {nv!s}")
    if not dry_run:
        for rid, fcs in by_item.items():
            for f, ov, nv in fcs:
                write_audit_log(conn, "decisions", rid, f, ov, nv, "xlsx_sync")
            set_clause = ", ".join(f"{f} = ?" for f, _, _ in fcs)
            values = [nv for _, _, nv in fcs] + [rid]
            conn.execute(f"UPDATE decisions SET {set_clause} WHERE id = ?", values)
    return total_changes, len(by_item)


# --------------------------------------------------------------------------- #
# 設定
# --------------------------------------------------------------------------- #
def load_report_config(config_path: Path) -> dict:
    if not config_path.exists():
        return {}
    with open(config_path, encoding="utf-8") as f:
        cfg = yaml.safe_load(f) or {}
    return cfg.get("report") or {}


# --------------------------------------------------------------------------- #
# メイン
# --------------------------------------------------------------------------- #
def main() -> None:
    parser = argparse.ArgumentParser(
        description="Box の進捗レポート XLSX を pm.db に反映する"
    )
    parser.add_argument("--db", default=str(DEFAULT_PM_DB),
                        help=f"pm.db のパス（デフォルト: {DEFAULT_PM_DB}）")
    parser.add_argument("--config", default=str(DEFAULT_CONFIG))
    parser.add_argument("--xlsx", default=None,
                        help="ローカル XLSX を直接読む（指定時は Box ダウンロードをスキップ）")
    parser.add_argument("--box-folder-id", default=None,
                        help="Box folder ID（未指定時は report.box_folder_id）")
    parser.add_argument("--filename", default=None,
                        help=f"Box 上のファイル名（デフォルト: {DEFAULT_FILENAME}）")
    add_dry_run_arg(parser)
    add_no_encrypt_arg(parser)
    args = parser.parse_args()

    log, _ = make_logger(None)
    db_path = Path(args.db)
    report_cfg = load_report_config(Path(args.config))
    folder_id = args.box_folder_id or report_cfg.get("box_folder_id")
    filename = args.filename or report_cfg.get("filename") or DEFAULT_FILENAME

    if args.xlsx:
        xlsx_path = Path(args.xlsx)
        if not xlsx_path.exists():
            log(f"[ERROR] ファイルが見つかりません: {xlsx_path}")
            sys.exit(1)
    else:
        if not folder_id:
            log("[INFO] box_folder_id が未設定のため XLSX 同期をスキップ")
            return
        with tempfile.TemporaryDirectory() as td:
            xlsx_path = fetch_xlsx_from_box(folder_id, filename, Path(td), log)
            if xlsx_path is None:
                return
            return _process_xlsx(xlsx_path, db_path, args, log)

    _process_xlsx(xlsx_path, db_path, args, log)


def _process_xlsx(xlsx_path: Path, db_path: Path, args, log) -> None:
    log(f"[INFO] XLSX        : {xlsx_path}")
    log(f"[INFO] pm.db       : {db_path}")

    # 議事録 DB (data/minutes/*.db) は pm.db とスキーマが異なるため対象外
    if "minutes" in str(db_path):
        log(f"[ERROR] --db に議事録 DB が指定されています: {db_path}")
        log("[ERROR] pm_xlsx_sync.py は pm.db 専用です。--db data/pm.db を指定してください。")
        sys.exit(1)

    wb = load_workbook(xlsx_path, data_only=True)
    ai_rows = _read_sheet(wb, SHEET_AI, AI_FIELD_MAP)
    dec_rows = _read_sheet(wb, SHEET_DEC, DEC_FIELD_MAP)
    log(f"[INFO] アクション編集候補: {len(ai_rows)}件 / 決定事項編集候補: {len(dec_rows)}件")

    conn = open_pm_db(db_path, no_encrypt=args.no_encrypt)
    try:
        ai_changes, ai_items = _apply_action_items(conn, ai_rows, args.dry_run, log)
        dec_changes, dec_items = _apply_decisions(conn, dec_rows, args.dry_run, log)
        if not args.dry_run:
            conn.commit()
    finally:
        conn.close()

    log(f"[INFO] action_items: {ai_changes} 変更 / {ai_items} 件")
    log(f"[INFO] decisions    : {dec_changes} 変更 / {dec_items} 件")
    if args.dry_run:
        log("[INFO] dry-run のため DB は変更していません")


if __name__ == "__main__":
    main()
